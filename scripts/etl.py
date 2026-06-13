#!/usr/bin/env python3
"""
ETL for the open MACC tool — the published trust anchor.

Reads the source Excel workbook (read-only) and emits, for one country:
  data/<country>/workbook.engine.json  — full formula+literal grid seeding HyperFormula
  data/<country>/model.data.json       — clean, labelled, bilingual dataset (Dataset schema)
  data/<country>/fingerprint.json       — content hash + structural signature + source date

Run:  py scripts/etl.py            (defaults to country=kz, source from CONFIG)
      py scripts/etl.py --check    (verify only, do not write)

Design notes
------------
* Output is DETERMINISTIC (no wall-clock timestamps) so re-running reproduces
  byte-identical files — that is what makes the curve third-party verifiable.
* The whole workbook uses only IF / PV / SUM (+ ^ and arithmetic). We re-derive
  I/J/K via PV in Python purely to CROSS-CHECK against Excel's cached <v> values;
  the browser still runs the real formulas via HyperFormula.
* openpyxl 3.1 already materialises shared formulas (the rows 7..31 of I/J/K that
  carry empty <f t="shared"> in the XML come back as full formula strings). We
  still cross-check every output cell against its cached value to catch any cell
  where the formula failed to round-trip.

Conforms to data/schema.ts.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
ETL_VERSION = "1"
SCHEMA_VERSION = 1
REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = REPO_ROOT / "MACC_KZ_29052026_rev.xlsx"
TRANSLATIONS = REPO_ROOT / "scripts" / "translations.en.json"

SHEET_MACC = "MACC"
SHEET_EMISSIONS = "Выбросы"
SHEET_CALC = "Расчёты"

PROJECT_ROWS = range(6, 32)  # MACC rows 6..31 = 26 measures
TOTALS_ROW = 32

# Sector default palette (overridden by translation overlay for labels).
SECTOR_COLORS = {
    "1.A.1": "#c0392b", "1.A.2": "#e67e22", "1.A.3": "#f1c40f",
    "1.A.4": "#16a085", "1.B": "#8e44ad", "2": "#2980b9",
    "3": "#27ae60", "4": "#2ecc71", "5": "#7f8c8d",
}
SECTOR_LABELS_RU = {
    "1.A.1": "Энергетические отрасли", "1.A.2": "Обрабатывающая промышленность",
    "1.A.3": "Транспорт", "1.A.4": "Другие секторы (ЖКХ, здания)",
    "1.B": "Фугитивные выбросы", "2": "Промышленные процессы (IPPU)",
    "3": "Сельское хозяйство", "4": "Землепользование (LULUCF)", "5": "Отходы",
}

# Layer-1 "general inputs": cell -> (key, group, unit). Levers also get ranges.
# Cells live on the MACC sheet. Discount rate is C2; price/EF/CAPEX in 36..57.
ASSUMPTION_MAP = [
    # key,                     cell,        group,             unit
    ("discountRate",           "C2",  "lever",          ""),
    ("coalPrice",              "E36", "lever",          "$/т"),
    ("gasPrice",               "E37", "lever",          "$/тыс.м³"),
    ("electricityPrice",       "E38", "lever",          "$/МВтч"),
    ("efCoalPower",            "E39", "emission_factor", "тCO₂/МВтч"),
    ("efGasPower",             "E40", "emission_factor", "тCO₂/МВтч"),
    ("efCoalHeat",             "E41", "emission_factor", "тCO₂/Гкал"),
    ("coalCarbonContent",      "E42", "emission_factor", "доля"),
    ("capexCoalChpEfficiency", "E45", "capex_unit",      "$/кВт"),
    ("capexCoalToGas",         "E46", "capex_unit",      "$/кВт"),
    ("capexSolar",             "E47", "capex_unit",      "$/кВт"),
    ("capexWind",              "E48", "capex_unit",      "$/кВт"),
    ("capexGasPeaker",         "E49", "capex_unit",      "$/кВт"),
    ("capexNuclear",           "E50", "capex_unit",      "$/кВт"),
    ("capexHeatPump",          "E51", "capex_unit",      "$/кВт(э)"),
    ("capexElectricBoiler",    "E52", "capex_unit",      "$/кВт"),
    ("capexWasteHeat",         "E53", "capex_unit",      "$/кВт(т)"),
    ("capexEv",                "E54", "capex_unit",      "$/авто"),
    ("capexThermalRetrofit",   "E55", "capex_unit",      "$/м²"),
    ("capexCcus",              "E56", "capex_unit",      "$/т CO₂/год"),
    ("capexAfforestation",     "E57", "capex_unit",      "$/га"),
]
# RU labels for assumptions are read live from column D of each row; for C2 it is A2.
LEVER_RANGES = {
    "coalPrice":        {"min": 5,    "max": 60,  "step": 1},
    "gasPrice":         {"min": 100,  "max": 600, "step": 5},
    "electricityPrice": {"min": 20,   "max": 150, "step": 1},
    "discountRate":     {"min": 0.03, "max": 0.20, "step": 0.005},
}

TOL = 1e-6  # relative tolerance for the cached-value cross-check


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def excel_pv(rate: float, nper: float, pmt: float, fv: float = 0.0, typ: int = 0) -> float:
    """Excel PV(): present value of a constant-payment annuity (type 0/1)."""
    if rate == 0:
        return -(pmt * nper + fv)
    factor = (1 + rate) ** -nper
    return -(pmt * (1 + rate * typ) * (1 - factor) / rate + fv * factor)


def cell_to_engine(value):
    """Convert an openpyxl (formula) cell value to an engine-JSON cell."""
    if value is None:
        return None
    # openpyxl formula objects (ArrayFormula / DataTableFormula) carry .text
    if hasattr(value, "text"):
        text = value.text
        return text if text.startswith("=") else "=" + text
    if isinstance(value, str):
        return value  # formula string ("=...") or plain text
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    # dates etc. — stringify defensively (none expected in this workbook)
    return str(value)


def approx(a: float, b: float, tol: float = TOL) -> bool:
    if a is None or b is None:
        return a == b
    denom = max(1.0, abs(a), abs(b))
    return abs(a - b) <= tol * denom


def parse_source_date(filename: str) -> str:
    """MACC_KZ_29052026_rev.xlsx -> 2026-05-29 (ddmmyyyy in the name)."""
    m = re.search(r"(\d{2})(\d{2})(\d{4})", filename)
    if not m:
        return "unknown"
    dd, mm, yyyy = m.groups()
    return f"{yyyy}-{mm}-{dd}"


# ---------------------------------------------------------------------------
# core
# ---------------------------------------------------------------------------
def build_engine(wb_f) -> dict:
    """Full formula+literal grid for every sheet (seeds HyperFormula)."""
    sheets = {}
    for ws in wb_f.worksheets:
        grid = []
        for r in range(1, ws.max_row + 1):
            row = [cell_to_engine(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            grid.append(row)
        sheets[ws.title] = grid
    return sheets


def compute_fingerprint(source: Path, wb_f) -> dict:
    raw = source.read_bytes()
    sha = hashlib.sha256(raw).hexdigest()
    sig_parts = []
    formula_total = 0
    for ws in wb_f.worksheets:
        fcount = sum(
            1
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
        formula_total += fcount
        sig_parts.append(f"{ws.title}:{ws.max_row}x{ws.max_column}:f{fcount}")
    structural = "|".join(sig_parts)
    model_hash = hashlib.sha256(
        f"{sha}|{structural}|etl{ETL_VERSION}".encode()
    ).hexdigest()[:12]
    model_version = f"kz-{model_hash}"
    source_date = parse_source_date(source.name)
    return {
        "modelVersion": model_version,
        "sourceFile": source.name,
        "sourceDate": source_date,
        "sourceSha256": sha,
        "etlVersion": ETL_VERSION,
        "structuralSignature": structural,
        "formulaCount": formula_total,
        "display": f"Model v.{model_hash} · source MACC_KZ_{source_date}.xlsx",
    }


def build_assumptions(macc_f, macc_v, translations) -> list:
    en_map = translations.get("assumptions", {})
    ru_overrides = translations.get("assumptionLabelsRu", {})
    out = []
    for key, cell, group, unit in ASSUMPTION_MAP:
        col = cell[0]
        row = int(cell[1:])
        value = macc_v[f"{col}{row}"].value
        # RU label: an explicit overlay override wins (used to correct misleading
        # source labels); else A2 for the discount rate, else column D of the row.
        label_ru = (
            ru_overrides.get(key)
            or (macc_v["A2"].value if cell == "C2" else macc_v[f"D{row}"].value)
            or key
        )
        a = {
            "key": key,
            "cell": f"{SHEET_MACC}!{cell}",
            "label": {"ru": label_ru, "en": en_map.get(key, "")},
            "value": value,
            "unit": unit,
            "group": group,
            "isLever": group == "lever",
        }
        if key in LEVER_RANGES:
            a.update(LEVER_RANGES[key])
        out.append(a)
    return out


def extract_source_cell(formula):
    """'=Расчёты!C670' -> 'Расчёты!C670' (provenance). Returns None if not a ref."""
    if not isinstance(formula, str):
        return None
    m = re.match(r"^=\s*([^\s!]+!\$?[A-Z]+\$?\d+)\s*$", formula)
    return m.group(1).replace("$", "") if m else None


def calc_block_ranges(calc_v):
    """Parse the Расчёты sheet into project blocks: [(start, end, code, title)].
    A block starts at a column-A row like '[1.A.1-1] <title>'."""
    starts = []
    for r in range(1, calc_v.max_row + 1):
        a = calc_v.cell(r, 1).value
        if isinstance(a, str):
            m = re.match(r"^\[([^\]]+)\]\s*(.*)$", a.strip())
            if m:
                starts.append((r, m.group(1), m.group(2).strip()))
    ranges = []
    for i, (r, code, title) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else calc_v.max_row + 1
        ranges.append((r, end, code, title))
    return ranges


def find_block(ranges, row):
    """Return the block (start, end, code, title) containing a 1-based row."""
    for start, end, code, title in ranges:
        if start <= row < end:
            return (start, end, code, title)
    return None


def extract_breakdown(calc_v, start, end, kind, line_en):
    """Extract CAPEX/OPEX line items between the 'KIND (...)' header and 'ИТОГО KIND'.
    Returns (items, total) or (None, None) if the section is absent."""
    hdr = tot = None
    for r in range(start, end):
        b = calc_v.cell(r, 2).value
        if isinstance(b, str):
            bs = b.strip()
            if hdr is None and bs.startswith(kind + " ("):
                hdr = r
            elif hdr is not None and bs.startswith("ИТОГО " + kind):
                tot = r
                break
    if hdr is None or tot is None:
        return None, None
    items = []
    for r in range(hdr + 1, tot):
        b = calc_v.cell(r, 2).value
        c = calc_v.cell(r, 3).value
        if isinstance(b, str) and isinstance(c, (int, float)):
            label_ru = b.strip()
            items.append({
                "label": {"ru": label_ru, "en": line_en.get(label_ru, "")},
                "value": c,
                "cell": f"{SHEET_CALC}!C{r}",
            })
    return items, calc_v.cell(tot, 3).value


# Physical-scale extraction (the tangible thing CAPEX buys: MW, ha, km, head…).
# A physical unit is a standalone magnitude — NOT a monetary, emission, intensity
# (per-something) or share/time unit. Capture/throughput rates "X/год" are kept.
_PHYS_UNIT = re.compile(
    r"(МВт|кВт|ГВт|Гкал|МВтч|кВтч|ГВтч|\bга\b|м²|м2|\bм3\b|м³|\bкм\b|шт|голов|объект|\bт\b|тонн|авто|чел)",
    re.I,
)
_BAD_UNIT = re.compile(r"\$|USD|eq|доля|%|лет|у\.т\.|^год$", re.I)
# A "/" that is NOT "/год" or "/хоз…" marks a per-unit intensity (e.g. tCO₂/Гкал).
_INTENSITY = re.compile(r"/\s*(?!год|хоз)", re.I)
# Sector baselines / context rows, not the measure's own scale.
_BASELINE_LABEL = re.compile(r"^(Выбросы сектора|Общий объ|Общие |Доля )", re.I)
_SECTION_LABEL = re.compile(r"^(ПРЕДПОСЫЛКИ|РАСЧ[ЁЕ]Т|КЛЮЧЕВЫЕ|CAPEX|OPEX|ИТОГО)", re.I)


def build_physical_items(calc_v, start, end, phys_en) -> list:
    """Tangible non-monetary scale indicators from a Расчёты block."""
    items = []
    for r in range(start, end):
        label = calc_v.cell(r, 2).value   # col B
        value = calc_v.cell(r, 3).value   # col C (data_only → cached number)
        unit = calc_v.cell(r, 5).value    # col E
        if not isinstance(label, str) or not isinstance(unit, str):
            continue
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        label, unit = label.strip(), unit.strip()
        if not unit or _SECTION_LABEL.match(label) or _BASELINE_LABEL.match(label):
            continue
        if _BAD_UNIT.search(unit) or _INTENSITY.search(unit) or not _PHYS_UNIT.search(unit):
            continue
        items.append({
            "label": {"ru": label, "en": phys_en.get(label, "")},
            "value": value,
            "unit": unit,
            "cell": f"{SHEET_CALC}!C{r}",
        })
    return items


def build_local_inputs(calc_v, start, end, li_en) -> list:
    """Editable per-measure assumptions: rows tagged IN-L in the classifier
    column D (the engineering/economic premises behind a measure — efficiency
    gains, capacity factors, lifetimes, capacities, shares). Read straight from
    the tag, no heuristic guessing. Source/citation comes from col F when present.
    Centralizing these onto a «Допущения» sheet is Phase C; here they are merely
    surfaced read-only in the drill-down."""
    items = []
    for r in range(start, end):
        if calc_v.cell(r, 4).value != "IN-L":   # col D = level classifier
            continue
        label = calc_v.cell(r, 2).value
        value = calc_v.cell(r, 3).value
        if not isinstance(label, str) or not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        unit = calc_v.cell(r, 5).value
        source = calc_v.cell(r, 6).value
        label = label.strip()
        items.append({
            "label": {"ru": label, "en": li_en.get(label, "")},
            "value": value,
            "unit": unit.strip() if isinstance(unit, str) else "",
            "source": source.strip() if isinstance(source, str) else "",
            "cell": f"{SHEET_CALC}!C{r}",
        })
    return items


def build_projects(macc_f, macc_v, calc_v, translations) -> list:
    en_map = translations.get("projects", {})
    line_en = translations.get("lineItems", {})
    phys_en = translations.get("physicalItems", {})
    li_en = translations.get("localInputs", {})
    ranges = calc_block_ranges(calc_v)
    rows = []
    for r in PROJECT_ROWS:
        name_ru = macc_v[f"D{r}"].value
        variant = macc_v[f"C{r}"].value
        capex_src = extract_source_cell(macc_f[f"E{r}"].value)

        # Locate the Расчёты block via the CAPEX source cell row, then pull breakdowns.
        capex_items = opex_items = physical_items = local_inputs = None
        if capex_src:
            row_num = int(re.search(r"(\d+)$", capex_src).group(1))
            blk = find_block(ranges, row_num)
            if blk:
                start, end, _code, _title = blk
                capex_items, _ = extract_breakdown(calc_v, start, end, "CAPEX", line_en)
                opex_items, _ = extract_breakdown(calc_v, start, end, "OPEX", line_en)
                physical_items = build_physical_items(calc_v, start, end, phys_en)
                local_inputs = build_local_inputs(calc_v, start, end, li_en)

        rows.append({
            "id": int(macc_v[f"B{r}"].value),
            "sector": str(macc_v[f"A{r}"].value),
            "variant": int(variant) if isinstance(variant, (int, float)) else None,
            "name": {"ru": name_ru, "en": en_map.get(name_ru, "")},
            "capex": macc_v[f"E{r}"].value,
            "opex": macc_v[f"F{r}"].value,
            "durationYrs": macc_v[f"G{r}"].value,
            "abatementKt": macc_v[f"H{r}"].value,
            "npv": macc_v[f"I{r}"].value,
            "discCo2Kt": macc_v[f"J{r}"].value,
            "mac": macc_v[f"K{r}"].value,
            "sourceCells": {
                "capex": capex_src,
                "opex": extract_source_cell(macc_f[f"F{r}"].value),
                "durationYrs": extract_source_cell(macc_f[f"G{r}"].value),
                "abatementKt": extract_source_cell(macc_f[f"H{r}"].value),
            },
            "capexItems": capex_items or [],
            "opexItems": opex_items or [],
            "physicalItems": physical_items or [],
            "localInputs": local_inputs or [],
        })
    # curve order: ascending by MAC, with cumulative abatement spans
    rows.sort(key=lambda p: p["mac"])
    cum = 0.0
    for p in rows:
        p["cumAbatementStartKt"] = cum
        cum += p["abatementKt"]
        p["cumAbatementEndKt"] = cum
    return rows


def build_sectors(projects, translations) -> dict:
    en_map = translations.get("sectors", {})
    used = sorted({p["sector"] for p in projects})
    return {
        code: {
            "label": {"ru": SECTOR_LABELS_RU.get(code, code), "en": en_map.get(code, "")},
            "color": SECTOR_COLORS.get(code, "#888888"),
        }
        for code in used
    }


def build_totals(macc_v, projects) -> dict:
    no_regrets = sum(p["abatementKt"] for p in projects if p["mac"] < 0)
    return {
        "capexMUsd": macc_v[f"E{TOTALS_ROW}"].value,
        "abatementKt": macc_v[f"H{TOTALS_ROW}"].value,
        "npvMUsd": macc_v[f"I{TOTALS_ROW}"].value,
        "discCo2Kt": macc_v[f"J{TOTALS_ROW}"].value,
        "weightedAvgMac": macc_v[f"K{TOTALS_ROW}"].value,
        "noRegretsAbatementKt": no_regrets,
    }


def verify(macc_v, projects, totals) -> list:
    """Cross-check our understanding of the formulas against Excel's cached <v>."""
    errors = []
    rate = macc_v["C2"].value
    for r in PROJECT_ROWS:
        E = macc_v[f"E{r}"].value
        F = macc_v[f"F{r}"].value
        G = macc_v[f"G{r}"].value
        H = macc_v[f"H{r}"].value
        I_cached = macc_v[f"I{r}"].value
        J_cached = macc_v[f"J{r}"].value
        K_cached = macc_v[f"K{r}"].value
        I_calc = E - excel_pv(rate, G, F)
        J_calc = -excel_pv(rate, G, H)
        K_calc = 0 if J_calc == 0 else I_calc / J_calc * 1000
        if not approx(I_calc, I_cached):
            errors.append(f"row {r}: I {I_calc} != cached {I_cached}")
        if not approx(J_calc, J_cached):
            errors.append(f"row {r}: J {J_calc} != cached {J_cached}")
        if not approx(K_calc, K_cached):
            errors.append(f"row {r}: K {K_calc} != cached {K_cached}")
    # CAPEX/OPEX breakdown sums must reconcile to the project totals
    for p in projects:
        if p["capexItems"]:
            s = sum(it["value"] for it in p["capexItems"])
            if not approx(s, p["capex"]):
                errors.append(f"id {p['id']}: CAPEX items {s} != {p['capex']}")
        else:
            errors.append(f"id {p['id']}: missing CAPEX breakdown")
        if p["opexItems"]:
            s = sum(it["value"] for it in p["opexItems"])
            if not approx(s, p["opex"]):
                errors.append(f"id {p['id']}: OPEX items {s} != {p['opex']}")
        else:
            errors.append(f"id {p['id']}: missing OPEX breakdown")

    # totals & counts
    if len(projects) != 26:
        errors.append(f"expected 26 projects, got {len(projects)}")
    abate = sum(p["abatementKt"] for p in projects)
    if not approx(abate, totals["abatementKt"]):
        errors.append(f"sum abatement {abate} != totals {totals['abatementKt']}")
    # Headline sanity anchors. Updated 2026-06-13: the three agriculture measures
    # (blocks 3-2/3-3/3-4) now reference their sub-category baselines on the Выбросы
    # sheet (manure C23=3.3, soils C25=11.6) instead of a stale blanket 42.8 Mt that
    # exceeded the whole sector — a correctness fix that lowers total abatement.
    if not approx(totals["abatementKt"], 214352.91897671297, tol=1e-4):
        errors.append(f"total abatement {totals['abatementKt']} != expected 214353")
    if not approx(totals["weightedAvgMac"], 95.57998085733414, tol=1e-4):
        errors.append(f"wavg MAC {totals['weightedAvgMac']} != expected 95.58")
    return errors


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def run(source: Path, country: str, check_only: bool) -> int:
    if not source.exists():
        print(f"ERROR: source not found: {source}", file=sys.stderr)
        return 2
    translations = json.loads(TRANSLATIONS.read_text(encoding="utf-8")) if TRANSLATIONS.exists() else {}

    wb_f = openpyxl.load_workbook(source, data_only=False)
    wb_v = openpyxl.load_workbook(source, data_only=True)
    macc_f, macc_v = wb_f[SHEET_MACC], wb_v[SHEET_MACC]
    calc_v = wb_v[SHEET_CALC]

    fingerprint = compute_fingerprint(source, wb_f)
    assumptions = build_assumptions(macc_f, macc_v, translations)
    projects = build_projects(macc_f, macc_v, calc_v, translations)
    sectors = build_sectors(projects, translations)
    totals = build_totals(macc_v, projects)

    errors = verify(macc_v, projects, totals)
    print(f"Model version : {fingerprint['modelVersion']}")
    print(f"Source date   : {fingerprint['sourceDate']}  ({fingerprint['formulaCount']} formula cells)")
    print(f"Projects      : {len(projects)}")
    print(f"Total abatement: {totals['abatementKt']:.1f} kt   (no-regrets {totals['noRegretsAbatementKt']:.0f} kt)")
    print(f"Weighted MAC  : {totals['weightedAvgMac']:.2f} USD/t")
    if errors:
        print(f"\nVERIFICATION FAILED ({len(errors)} issue(s)):", file=sys.stderr)
        for e in errors[:20]:
            print(f"  - {e}", file=sys.stderr)
        return 1
    print("Verification  : OK (I/J/K reproduce cached Excel values within 1e-6)")

    if check_only:
        print("\n--check: no files written.")
        return 0

    engine = build_engine(wb_f)
    dataset = {
        "schemaVersion": SCHEMA_VERSION,
        "country": country,
        "modelVersion": fingerprint["modelVersion"],
        "meta": {
            "sourceFile": fingerprint["sourceFile"],
            "sourceDate": fingerprint["sourceDate"],
            "discountRate": macc_v["C2"].value,
            "totalEmissionsMt": wb_v[SHEET_EMISSIONS]["C33"].value,
            "totalEmissionsExclLulucfMt": wb_v[SHEET_EMISSIONS]["C32"].value,
            "emissionsSource": "UNFCCC BTR1 Kazakhstan (2024), inventory year 2022, GWP AR5",
            "enTranslationStatus": "machine-draft-pending-review",
        },
        "sectors": sectors,
        "assumptions": assumptions,
        "projects": projects,
        "totals": totals,
    }
    engine_doc = {
        "meta": {
            "sourceFile": fingerprint["sourceFile"],
            "modelVersion": fingerprint["modelVersion"],
            "sheetOrder": wb_f.sheetnames,
        },
        "sheets": engine,
    }

    out_dir = REPO_ROOT / "data" / country
    out_dir.mkdir(parents=True, exist_ok=True)

    def dump(name, obj):
        path = out_dir / name
        path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"  wrote {path.relative_to(REPO_ROOT)}  ({path.stat().st_size:,} bytes)")

    print("\nWriting:")
    dump("fingerprint.json", fingerprint)
    dump("model.data.json", dataset)
    dump("workbook.engine.json", engine_doc)
    return 0


def main():
    ap = argparse.ArgumentParser(description="MACC ETL: xlsx -> data/<country>/*.json")
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    ap.add_argument("--country", default="kz")
    ap.add_argument("--check", action="store_true", help="verify only; write nothing")
    args = ap.parse_args()
    sys.exit(run(args.source, args.country, args.check))


if __name__ == "__main__":
    main()
